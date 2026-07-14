"""
Generate WATT-IF Test Cases Excel file matching the game-dev template format.
Each module/area gets its own sheet with header block, pre-conditions, and test case table.

Run: pip install openpyxl   (if not installed)
Then: python generate_test_cases_excel.py
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def parse_test_cases_from_md(filepath):
    """Parse a markdown test case file into structured data."""
    text = Path(filepath).read_text(encoding="utf-8")

    # Extract metadata
    module_match = re.search(r"\*\*(?:Module|Area):\*\*\s*(.+)", text)
    module_name = module_match.group(1).strip() if module_match else ""

    precond_match = re.search(r"\*\*Pre-condition[s]?:\*\*\s*(.+?)(?:\n\*\*|\n---)", text, re.DOTALL)
    precondition = precond_match.group(1).strip() if precond_match else ""

    dep_match = re.search(r"\*\*Dependencies:\*\*\s*(.+?)(?:\n\*\*|\n---)", text, re.DOTALL)
    dependencies = dep_match.group(1).strip() if dep_match else "N/A"

    prio_match = re.search(r"\*\*Test Priority:\*\*\s*(.+)", text)
    priority = prio_match.group(1).strip() if prio_match else "Medium"

    # Parse individual test cases
    test_cases = []
    # Split by test case headers (### PREFIX-NN: Title)
    tc_pattern = r"###\s+([\w-]+):\s*(.+?)(?=\n###|\Z)"
    tc_blocks = re.finditer(tc_pattern, text, re.DOTALL)

    for match in tc_blocks:
        tc_id = match.group(1).strip()
        rest = match.group(2).strip()

        # Extract title (first line)
        lines = rest.split("\n")
        title = lines[0].strip() if lines else ""

        # Extract fields
        summary = extract_field(rest, "Summary")
        steps = extract_field(rest, "Test Steps")
        test_data = extract_field(rest, "Test Data")
        expected = extract_field(rest, "Expected Result")
        postcond = extract_field(rest, "Post-condition")
        actual = extract_field(rest, "Actual Result")
        status = extract_field(rest, "Status")
        notes = extract_field(rest, "Notes")

        # Clean status
        if "Not Run" in status:
            status = "Not Run"
        elif not status:
            status = "Not Run"

        test_cases.append({
            "id": tc_id,
            "title": title,
            "summary": summary,
            "steps": steps,
            "test_data": test_data,
            "expected": expected,
            "postcondition": postcond,
            "actual": actual if actual and "to be filled" not in actual else "",
            "status": status,
            "notes": notes,
        })

    return {
        "module": module_name,
        "precondition": precondition,
        "dependencies": dependencies,
        "priority": priority,
        "test_cases": test_cases,
    }


def extract_field(text, field_name):
    """Extract a field value from markdown test case block."""
    pattern = rf"\*\*{re.escape(field_name)}:\*\*\s*(.+?)(?=\n\*\*|\n---|\n###|\Z)"
    match = re.search(pattern, text, re.DOTALL)
    if match:
        val = match.group(1).strip()
        # Clean markdown formatting
        val = re.sub(r"\*\*", "", val)
        val = val.replace("⬜ ", "")
        return val
    return ""


def create_sheet(wb, sheet_name, data):
    """Create a formatted sheet matching the game-dev template."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31 char limit

    # Styles
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=12)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    # Row 1-2: Project info header
    ws.merge_cells("A1:B1")
    ws["A1"] = "Project Name:"
    ws["A1"].font = header_font
    ws["C1"] = "WATT-IF"
    ws["C1"].font = title_font
    ws["F1"] = "Test Designed by:"
    ws["F1"].font = header_font
    ws["G1"] = "QA Team"

    ws.merge_cells("A2:B2")
    ws["A2"] = "Module Name:"
    ws["A2"].font = header_font
    ws["C2"] = data["module"][:50]
    ws["F2"] = "Test Designed Date:"
    ws["F2"].font = header_font
    ws["G2"] = "July 2026"

    ws.merge_cells("A3:B3")
    ws["A3"] = "Release Version:"
    ws["A3"].font = header_font
    ws["C3"] = "v1.0"
    ws["F3"] = "Test Executed by:"
    ws["F3"].font = header_font
    ws["G3"] = ""
    ws["A4"] = ""
    ws["F4"] = "Test Execution date:"
    ws["F4"].font = header_font
    ws["G4"] = ""

    # Row 6-8: Pre-conditions
    ws["A6"] = "Pre-condition"
    ws["A6"].font = header_font
    ws.merge_cells("B6:J6")
    ws["B6"] = data["precondition"]
    ws["B6"].alignment = wrap_align

    ws["A7"] = "Dependencies"
    ws["A7"].font = header_font
    ws.merge_cells("B7:J7")
    ws["B7"] = data["dependencies"]

    ws["A8"] = "Test Priority"
    ws["A8"].font = header_font
    ws["B8"] = data["priority"]

    # Row 10: Table headers
    headers = [
        "Test Case #", "Test Title", "Test Summary", "Test Steps",
        "Test Data", "Expected Result", "Post-condition",
        "Actual Result", "Status", "Notes"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Row 11+: Test case data
    row = 11
    for tc in data["test_cases"]:
        ws.cell(row=row, column=1, value=tc["id"]).border = thin_border
        ws.cell(row=row, column=2, value=tc["title"]).border = thin_border
        ws.cell(row=row, column=3, value=tc["summary"]).border = thin_border
        ws.cell(row=row, column=4, value=tc["steps"]).border = thin_border
        ws.cell(row=row, column=5, value=tc["test_data"]).border = thin_border
        ws.cell(row=row, column=6, value=tc["expected"]).border = thin_border
        ws.cell(row=row, column=7, value=tc["postcondition"]).border = thin_border
        ws.cell(row=row, column=8, value=tc["actual"]).border = thin_border
        ws.cell(row=row, column=9, value=tc["status"]).border = thin_border
        ws.cell(row=row, column=10, value=tc["notes"]).border = thin_border

        # Apply wrap text to all cells in the row
        for col in range(1, 11):
            ws.cell(row=row, column=col).alignment = wrap_align

        row += 1

    # Set column widths
    col_widths = [12, 30, 40, 50, 30, 40, 30, 20, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    doc_dir = Path(__file__).parent

    # Define all test case files and their sheet names
    files = [
        ("TC_ACT_AccountSystem.md", "Account System"),
        ("TC_DM_DataManagement.md", "Data Management"),
        ("TC_FD_ForecastDashboard.md", "Forecast Dashboard"),
        ("TC_CHT_Chat.md", "Chat Assistant"),
        ("TC_PCT_PriceCalculator.md", "Price Calculator"),
        ("TC_SET_Settings.md", "Settings"),
        ("TC_SYS_SystemInfrastructure.md", "System Infrastructure"),
        ("TC_SEC_Security.md", "Security"),
        ("TC_PERF_Performance.md", "Performance"),
        ("TC_AIR_AIRobustness.md", "AI Robustness"),
        ("TC_BRWS_BrowserCompatibility.md", "Browser Compatibility"),
        ("TC_DEV_DeviceCompatibility.md", "Device Compatibility"),
        ("TC_A11Y_Accessibility.md", "Accessibility"),
    ]

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for filename, sheet_name in files:
        filepath = doc_dir / filename
        if not filepath.exists():
            print(f"WARNING: {filename} not found, skipping.")
            continue

        print(f"Processing {filename}...")
        data = parse_test_cases_from_md(filepath)
        create_sheet(wb, sheet_name, data)
        print(f"  -> {len(data['test_cases'])} test cases found.")

    output_path = doc_dir / "WATT-IF_Test_Cases.xlsx"
    wb.save(output_path)
    print(f"\nExcel file saved: {output_path}")


if __name__ == "__main__":
    main()
